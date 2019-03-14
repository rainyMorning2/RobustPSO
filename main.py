import benchmarks
import PSO
import Robust_PSO
import openpyxl
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import matplotlib.cm as cm
from matplotlib.colors import LogNorm


def run():

    # number of particles
    N = 50
    # dimensions
    D = 2
    # iteration times
    T = 50
    # number of random particles
    m = 5
    # bias
    e = 0.5
    # number of cal
    Time = 3

    benchmark = [benchmarks.testRobust(N, 2),
                 benchmarks.f1(N,20),
                 benchmarks.f2(N,20),
                 benchmarks.f3(N,20),
                 benchmarks.f5(N,20),
                 benchmarks.f6(N,2)]

    '算法数'
    num = 12



    result = []
    names = []

    for index in range(0, 6):
        # index of benchmarks

        wbk = openpyxl.Workbook()

        sheets = [wbk.create_sheet('data', 0), wbk.create_sheet('GPSO', 1), wbk.create_sheet('LPSO', 2),
                  wbk.create_sheet('SFPSO', 3), wbk.create_sheet('LFIPSO', 4), wbk.create_sheet("SFIPSO", 5),
                  wbk.create_sheet('SIPSO', 6), wbk.create_sheet('rGPSO', 7),
                  wbk.create_sheet('rLPSO', 8),wbk.create_sheet('rSFPSO', 9), wbk.create_sheet('rLFIPSO', 10),
                  wbk.create_sheet("rSFIPSO", 11),wbk.create_sheet('rSIPSO', 12)]

        name = str(type(benchmark[index]))[19:-2]
        # name = "testBenchmark"
        times = 0

        if index == 0:
            e = 0.5
        elif index == 5:
            e = 3
        else:
            e = 0.2

        biasPos = []
        biasFit = []
        temp = []
        pos = [np.zeros([Time, benchmark[index].D])]*num
        print(pos)
        while times < Time:

            print(times)
            data = []
            data.append(PSO.GPSO(benchmark[index], T))
            data.append(PSO.LPSO(benchmark[index], T))
            data.append(PSO.SFPSO(benchmark[index], T))
            data.append(PSO.LFIPSO(benchmark[index], T))
            data.append(PSO.SFIPSO(benchmark[index], T))
            data.append(PSO.SIPSO(benchmark[index], T))
            data.append(Robust_PSO.GPSO(benchmark[index], T, m, e))
            data.append(Robust_PSO.LPSO(benchmark[index], T, m, e))
            data.append(Robust_PSO.SFPSO(benchmark[index], T, m, e))
            data.append(Robust_PSO.LFIPSO(benchmark[index], T, m, e))
            data.append(Robust_PSO.SFIPSO(benchmark[index], T, m, e))
            data.append(Robust_PSO.SIPSO(benchmark[index], T, m, e))

            for i in range(0, num):
                for j in range(0, data[i][0].__len__()):
                    sheets[i + 1].cell(row=j + 2, column=times + 2).value = data[i][0][j]

                c=0
                for k in data[i][1][0]:
                    sheets[i+1].cell(row=data[i][0].__len__()+2+c,column=times+2).value = k  # x
                    pos[i][times][c] = k
                    c += 1


            times += 1

        for k in range(0,num):
            maxrow = sheets[k+1].max_row
            for i in range(0, Time):
                biasPos.append(np.zeros([m, benchmark[index].D]))
                biasFit.append(np.zeros(m))
                for j in range(0, m):
                    biasPos[i][j] = pos[k][i] + (np.random.rand(benchmark[index].D) * 2 * e - e)
                PSO.checkBoundary(biasPos[i], benchmark[index].low, benchmark[index].high)
                biasFit[i] = benchmark[index].getValue(biasPos[i])

                sheets[k + 1].cell(row=maxrow+1, column=i + 2).value = max(biasFit[i])

        avg = []
        selected = []

        for i in range(0, num):

            avg.append([])
            selected.append([])

            for j in range(0, Time):
                sheets[i + 1].cell(row=1, column=j + 2).value = j + 1

            # for j in range(0, data[i][0].__len__()):
            #     sheets[i + 1].cell(row=j + 2, column=1).value = (j + 1) * N

            for col in range(2, sheets[i + 1].max_column + 1):
                # if sheets[i + 1].cell(row=sheets[i + 1].max_row, column=col).value < benchmarks[index].goal:
                selected[i].append(col)

            for row in range(2, sheets[i + 1].max_row+1):
                avg[i].append(0)
                for col in selected[i]:
                    avg[i][row - 2] += sheets[i + 1].cell(row=row, column=col).value
                if selected[i].__len__() != 0:
                    avg[i][row - 2] /= selected[i].__len__()

        for i in range(0, num):
            for row in range(2, avg[i].__len__()+2):
                # sheets[0].cell(row=row, column=1).value = (row - 1) * N
                sheets[0].cell(row=row, column=i + 2).value = avg[i][row - 2]

        sheets[0].cell(row=1, column=2).value = "GPSO"
        sheets[0].cell(row=1, column=3).value = "LSO"
        sheets[0].cell(row=1, column=4).value = "SFPSO"
        sheets[0].cell(row=1, column=5).value = "LFIPSO"
        sheets[0].cell(row=1, column=6).value = "SFIPSO"
        sheets[0].cell(row=1, column=7).value = "SIPSO"
        sheets[0].cell(row=1, column=8).value = "rGPSO"
        sheets[0].cell(row=1, column=9).value = "rLPSO"
        sheets[0].cell(row=1, column=10).value = "rSFPSO"
        sheets[0].cell(row=1, column=11).value = "rLFIPSO"
        sheets[0].cell(row=1, column=12).value = "rSFIPSO"
        sheets[0].cell(row=1, column=13).value = "rSIPSO"

        # '成功率'
        # sheets[0].cell(row=1, column=9).value = selected[0].__len__() / Time
        # sheets[0].cell(row=1, column=10).value = selected[1].__len__() / Time
        # sheets[0].cell(row=1, column=11).value = selected[2].__len__() / Time
        # sheets[0].cell(row=1, column=12).value = selected[3].__len__() / Time
        # sheets[0].cell(row=1, column=13).value = selected[4].__len__() / Time

        wbk.save('.//result//' + name + '.xlsx')

        # bingoRate.append([])
        # result.append([])
        # names.append(name)
        # for i in range(0, num):
        #     bingoRate[index].append(selected[i].__len__() / Time)
        #     result[index].append(avg[i][avg[i].__len__() - 1])

    # wbk = openpyxl.Workbook()
    # sheet = wbk.create_sheet('result', 0)
    #
    # for j in range(0, names.__len__()):
    #     sheet.cell(row=j + 2, column=1).value = name[j]
    #     for i in range(0, num):
    #         sheet.cell(row=j + 2, column=i + 2).value = result[j][i]
    #         sheet.cell(row=j + 2, column=i + num + 2).value = bingoRate[j][i]
    #
    # wbk.save('.//result//' + 'result' + '.xlsx')


def test():

    # number of particles
    N = 50
    # dimensions
    D = 20
    # iteration times
    T = 500
    # number of random particles
    m = 8
    # bias
    e = 0.2

    b = benchmarks.f2(N, D)

    # t = np.zeros([1,2])
    # t[0][0]=0.4
    # t[0][1]=0.2
    # print(b.getValue(t)+35)
    #
    # t[0][0] = 2
    # t[0][1] = 1
    # print(b.getValue(t)+35)
    #
    # t[0][0] = 2.8
    # t[0][1] = 4.0
    # print(b.getValue(t)+35)
    #
    # t[0][0] = 1
    # t[0][1] = 2.5
    # print(b.getValue(t) + 35)

    print(Robust_PSO.GPSO(b, T, m, e)[0][T-1])
    print(Robust_PSO.LPSO(b, T, m, e)[0][T-1])
    print(Robust_PSO.SFPSO(b, T, m, e)[0][T-1])
    print(Robust_PSO.LFIPSO(b, T, m, e)[0][T-1])
    print(Robust_PSO.SFIPSO(b, T, m, e)[0][T-1])
    print(Robust_PSO.SIPSO(b, T, m, e)[0][T-1])

    # print(Robust_PSO.GPSO(b, T, m, e)[0])
    # print(Robust_PSO.LPSO(b, T, m, e)[0])
    # print(Robust_PSO.SFPSO(b, T, m, e)[0])
    # print(Robust_PSO.LFIPSO(b, T, m, e)[0])
    # print(Robust_PSO.SFIPSO(b, T, m, e)[0])
    # print(Robust_PSO.SIPSO(b, T, m, e)[0])
def draw(filename):

    wb = load_workbook('.//result//' + filename + '.xlsx', read_only=True)
    sheet = wb.get_sheet_by_name("SFPSO")

    point_x = []
    point_y = []

    for i in range(2,sheet.max_column+1):
        point_x.append(sheet.cell(row=sheet.max_row-1, column=i).value)
        point_y.append(sheet.cell(row=sheet.max_row, column=i).value)

    div = 200
    N = div*div
    D = 2

    b = benchmarks.testRobust(N, D)
    x, y = np.linspace(-0.6, 3.1, div), np.linspace(4.3, -0.3, div)
    position = np.zeros([N, D])

    index1 = 0
    index2 = 0
    for i in range(0, N):
        position[i][0] = x[index1]
        position[i][1] = y[index2]
        index1 = index1 + 1
        if index1 == div:
            index1 = 0
            index2 = index2 + 1

    z = (b.getValue(position)).reshape(div, div)
    plt.imshow(z+35, extent=(np.amin(x), np.amax(x), np.amin(y), np.amax(y)), cmap=cm.hot, norm=LogNorm())
    plt.colorbar()
    plt.plot(point_x, point_y, 'bo', marker='.', ms=5)
    plt.show()


if __name__ == '__main__':
    run()
    # draw("testBenchmark")
    # test()